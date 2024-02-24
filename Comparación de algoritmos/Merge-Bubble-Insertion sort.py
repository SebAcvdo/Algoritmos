""" 
Acevedo Juárez Sebastián
Dra. Miriam Pescador Rojas
Análisis y diseño de algoritmos

        Con la librería openpyxl nos permitimos abrir una hoja de cálculo, 
        donde serán registrados los tiempos de cada algoritmo en microsegundos.

        Instrucciones para instalar la librería: 
            En la terminal de Visual Studio:
                --> pip install openpyxl
                --> python -m pip install openpyxl
            En el launcher de Python (Windows): 
                --> py -m pip install openpyxl

"""

import random, timeit
import openpyxl


def insertionSort(A,n):
    for i in range(1,n):
        j = i-1
        while j >= 0 and A[j] > A[j+1]:
                A[j + 1],A[j] = A[j],A[j+1]
                j -= 1
    return A


def bubbleSort(A,n):
    while True:
        swap = False
        for i in range(1,n):
            if A[i-1] > A[i]:
                A[i-1],A[i] = A[i], A[i-1]
                swap = True
        n -= 1
        if swap == False: break
    return A


def mergeSort(A):
    if len(A) == 1:
        return A
    izq = mergeSort(A[:len(A)//2])
    der = mergeSort(A[len(A)//2:])
    return merge(izq,der)
    

def merge(izq, der):
    temp = []
    i = 0
    j = 0
    while i < len(izq) and j< len(der):
        if izq[i] < der[j]:
            temp.append(izq[i])
            i += 1
        else:
            temp.append(der[j])
            j += 1
    if i < len(izq):
        while i < len(izq):
            temp.append(izq[i])
            i += 1
    if j < len(der):
        while j < len(der):
            temp.append(der[j])
            j += 1
    return temp


#Funcion que toma un arreglo de números ordenados, y la convierte en un arreglo que satisface el peor caso para Merge Sort
def peorCaso(A):
    if len(A) <= 1:
        return A
    if len(A) == 2:
        return A[::-1]
    
    izq = []
    der = []

    for i in range(0,len(A),2):
        izq.append(A[i])

    for i in range(1,len(A),2):
        der.append(A[i])

    temp = peorCaso(izq) + peorCaso(der)
    return temp


#Abre el documento.
wb = openpyxl.load_workbook("pruebas.xlsx")
#Selecciona la hoja en la cuál escribir.
hoja = wb['Ejec2']

#Insertion Sort
print("==============================INSERTION SORT:==============================")
for i in range(1,11):
    A=[]
    n = i*1000
    print("\t\t\tPara n = ", n)

    #Lista A, de N numeros, ordenada. (Mejor caso para Bubble, Insertion y Merge)
    for j in range(1,n+1):A.append(j)

    #Lista A invertida (Peor caso para Insertion y Bubble)
    B= A[::-1]
    
    #Lista A desordenada (Caso promedio)
    C = random.sample(A,len(A))  

    #Mejor Caso
    print("Ejecutando: Insertion Mejor Caso", end=" ")
    tiempo1IS = round((timeit.timeit(lambda: insertionSort(A,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=3,value=tiempo1IS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo1IS)
    
    #Caso Promedio
    print("Ejecutando: Insertion Caso Promedio", end=" ")
    tiempo2IS = round((timeit.timeit(lambda: insertionSort(C,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=6,value=tiempo2IS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo2IS)
    
    #Peor Caso
    print("Ejecutando: Insertion Peor Caso", end=" ")
    tiempo3IS = round((timeit.timeit(lambda: insertionSort(B,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=9,value=tiempo3IS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo3IS)

#Bubble Sort    
print("==============================BUBBLE SORT:==============================")
for i in range(1,11):
    A=[]
    n = i*1000
    print("\t\t\tPara n = ", n)

    #Lista A, de N numeros, ordenada. (Mejor caso para Bubble, Insertion y Merge)
    for j in range(1,n+1):A.append(j)

    #Lista A invertida (Peor caso para Insertion y Bubble)
    B= A[::-1]
    
    #Lista A desordenada (Caso promedio)
    C = random.sample(A,len(A))  

    #Mejor Caso
    print(f"Ejecutando: Bubble Mejor Caso", end=" ")
    tiempo1BS = round((timeit.timeit(lambda: bubbleSort(A,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=2,value=tiempo1BS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo1BS)
    
    #Caso Promedio
    print("Ejecutando: Bubble Caso Promedio", end=" ")
    tiempo2BS = round((timeit.timeit(lambda: bubbleSort(C,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=5,value=tiempo2BS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo2BS)
    
    #Peor Caso
    print("Ejecutando: Bubble Peor Caso", end=" ")
    tiempo3BS = round((timeit.timeit(lambda: bubbleSort(B,n) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=8,value=tiempo3BS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo3BS)

#Merge Sort
print("==============================MERGE SORT:==============================")
for i in range(1,11):
    A=[]
    n = i*1000
    print("\t\t\tPara n = ", n)

    #Lista A, de N numeros, ordenada. (Mejor caso para Bubble, Insertion y Merge)
    for j in range(1,n+1):A.append(j)
    
    #Lista A desordenada (Caso promedio)
    C = random.sample(A,len(A))  
    #Lista A desordenada de cierta manera en la función peorCaso(Peor Caso Merge)
    E= peorCaso(A)

    #Mejor Caso
    print("Ejecutando: Merge Mejor Caso", end=" ")
    tiempo1MS = round((timeit.timeit(lambda: mergeSort(A) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=4,value=tiempo1MS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo1MS)
    
    #Caso Promedio
    print("Ejecutando: Merge Caso Promedio", end=" ")
    tiempo2MS = round((timeit.timeit(lambda: mergeSort(C) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=7,value=tiempo2MS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo2MS)
    
    #Peor Caso
    print("Ejecutando: Merge Peor Caso", end=" ")
    tiempo3MS = round((timeit.timeit(lambda: mergeSort(E) , number=1)*100000), 3)
    hoja.cell(row=(i+3),column=10,value=tiempo3MS)
    print("=> EJECUCIÓN FINALIZADA tiempo =", tiempo3MS)

print("===================================================")

wb.save("pruebas.xlsx")
