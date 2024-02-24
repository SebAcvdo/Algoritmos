""" 
Acevedo Juárez Sebastián
Dra. Miriam Pescador Rojas
Análisis y diseño de algoritmos

        Con la librería openpyxl nos permitimos abrir una hoja de cálculo, 
        donde serán registrados los tiempos de cada algoritmo en microsegundos.

        Instrucciones para instalar la librería: 
            En la terminal de Visual Studio:
                --> pip install openpyxl
                --> python -m pip install openpyxl
            En el launcher de Python (Windows): 
                --> py -m pip install openpyxl

"""
import random, timeit

import openpyxl


#Abre el documento.
wb = openpyxl.load_workbook("pruebas.xlsx")
#Selecciona la hoja en la cuál escribir.
hoja = wb['Ejec1']


#Iterativo.
def Maximo1(L,n):       
    max = L[0]
    for i in range(n):
        if L[i] > max:
            max = L[i]
    return max

#Recursivo.
def Maximo2(L, n):      
    if n == 1: 
        return L[0]
    mitad = n//2
    Lizq = L[0:mitad]
    Lder = L[mitad:]
    izq = Maximo2(Lizq,mitad)
    der = Maximo2(Lder,n-mitad)
    if izq > der: return izq 
    else: return der

for i in range (2,52,2):  
    #Obtiene el número de elementos en base a i.
    n = i*100    
    print("n = ", n)

    #Genera arreglo(Numeros entre 0 y 100,000)
    L = [random.randint(0,100000) for j in range(n)]   

    #Toma el tiempo del algoritmo iterativo.
    tiempo1 = round((timeit.timeit(lambda: Maximo1(L,n) , number=1)*1000000), 3)

    #Toma el tiempo del algoritmo recusrivo.
    tiempo2 = round((timeit.timeit(lambda: Maximo2(L,n) , number=1)*1000000), 3)

    print("Tiempo iterativo =", tiempo1, " || Tiempo recursivo: ",tiempo2)
    print("========================================================")


    #Coloca los resultados en la Hoja de Cálculo.
    hoja.cell(row=(2+i/2),column=2,value=tiempo1)
    hoja.cell(row=(2+i/2),column=3,value=tiempo2)

#Guarda el archivo.
wb.save("pruebas.xlsx")

  


